# -*- coding: utf-8 -*-
"""
build_analysis.py  –  เพิ่มชีต เงินเข้าบริษัท + ใบปะหน้า SHOPEE ลงใน input_combined.xlsx
ใช้ได้ทั้งแบบ import (run()) และ command-line (python build_analysis.py INPUT OUTPUT)
"""
import sys, shutil, re as _re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins


SP_SHEETS = ['SPจ่ายMAFIA','SPจ่ายTRC','SPจ่ายUTOPIA','SPจ่ายFREEROAD','SPจ่ายWORKFORCE']

COST_URL = "https://docs.google.com/spreadsheets/d/1whwXYiUyDC0uo6oaudr_QqzCV6UFOWd0mL4Y2El9JRY/edit?usp=sharing"
NOTE_URL = "https://docs.google.com/spreadsheets/d/1p4-DX-Sq-Mvo_FNAkYtGMkdEBvPDCuPM_sZ_b-_Tm-Y/edit?usp=sharing"


def run(INPUT, OUTPUT):
    print('Reading Data ERP...')
    wb0 = load_workbook(INPUT, data_only=True)
    ws0 = wb0['Data ERP']
    erp_headers = [c.value for c in ws0[1]]
    N = len(erp_headers)

    def find_idx(kws):
        for i,h in enumerate(erp_headers):
            for k in kws:
                if k in str(h): return i
        return None

    shop_i  = find_idx(['พนักงานขาย'])
    order_i = find_idx(['เลขสั่งซื้อ Shopee (copy)','เลขสั่งซื้อ Shopee copy']) or find_idx(['เลขสั่งซื้อ Shopee'])
    qty_i   = find_idx(['จำนวนที่สั่ง'])
    sku_i   = find_idx(['รหัสสินค้า'])

    F   = get_column_letter(order_i + 1)
    Y   = get_column_letter(shop_i  + 1)
    QT  = get_column_letter(qty_i   + 1) if qty_i is not None else 'O'
    SKU = get_column_letter(sku_i   + 1) if sku_i is not None else 'I'
    print(f'  order={F}  shop={Y}  qty={QT}  sku={SKU}')

    rows = [list(r) for r in ws0.iter_rows(min_row=2, values_only=True)
            if str(r[shop_i] or '').startswith('SHOPEE-')]
    print(f'  SHOPEE rows: {len(rows)}')
    wb0.close()

    wb1 = load_workbook(INPUT, data_only=True)
    sp_hdrs = {str(c.value).strip(): get_column_letter(c.column)
               for c in wb1['Data SHOPEE'][1] if c.value}
    SP_MR = wb1['Data SHOPEE'].max_row

    def fsp(kws):
        for h,c in sp_hdrs.items():
            for k in kws:
                if k in h: return c
        return None

    SP_ORDER = fsp(['หมายเลขคำสั่งซื้อ'])
    SP_PRICE = fsp(['ราคาขาย'])
    SP_AP    = fsp(['Transaction Fee'])
    SP_AO    = fsp(['ค่าคอมมิชชั่น'])
    SP_AU    = fsp(['ค่าบริการ'])
    SP_AS    = fsp(['Shopee ออกให้'])
    SP_AR    = fsp(['ค่าจัดส่งที่ชำระโดยผู้ซื้อ'])
    print(f'  ORDER={SP_ORDER} PRICE={SP_PRICE} AP={SP_AP} AO={SP_AO} AU={SP_AU} AS={SP_AS} AR={SP_AR}')
    wb1.close()

    if INPUT != OUTPUT:
        shutil.copy2(INPUT, OUTPUT)
    wb = load_workbook(INPUT if INPUT == OUTPUT else OUTPUT)
    sp_mr = {s: max(wb[s].max_row, 2) for s in SP_SHEETS if s in wb.sheetnames}

    if 'เงินเข้าบริษัท' in wb.sheetnames: del wb['เงินเข้าบริษัท']
    ws = wb.create_sheet('เงินเข้าบริษัท', 0)

    FHDRS = ['ยอดขาย Y','ค่าธรรมเนียมการจ่าย AP','ค่าคอมมิชชั่น AO',
             'ค่าบริการ AU','รายการปรับปรุงจาก Shopee','ค่าธรรมเนียมการจ่ายนอกรอบ',
             'ค่าจัดส่ง Shopee ออกให้ AS','ค่าจัดส่งที่ชำระโดยผู้ซื้อ AR',
             'ค่าจัดส่งผู้ขาย','เงินเข้าบริษัท',
             'ต้นทุนสินค้า/ต่อชิ้น','ต้นทุนสินค้ารวม',
             'กำไร (บาท)','กำไร (%)','หมายเหตุ','สาเหตุติดลบ']

    all_hdrs = list(erp_headers) + FHDRS
    for c,h in enumerate(all_hdrs, 1):
        ws.cell(1,c).value = h or ''

    CAA=N+1;  CAB=N+2;  CAC=N+3;  CAD=N+4;  CAE=N+5
    CAF=N+6;  CAG=N+7;  CAH=N+8;  CAI=N+9;  CAJ=N+10
    CAK=N+11; CAL=N+12; CAM=N+13; CAN=N+14; CAO=N+15; CAP=N+16

    L = {c: get_column_letter(c) for c in range(CAA, CAP+1)}
    AJ = L[CAJ]; AA = L[CAA]; AH = L[CAH]

    CE = max(len(rows)+200, 1000)
    SE = max(SP_MR+100, 3011)

    def cf(r):
        return 'COUNTIF(${F}$2:${F}${e},{F}{r})'.format(F=F, e=CE, r=r)

    def sp_core(col, r):
        return ("SUMIFS('Data SHOPEE'!${c}$2:${c}${e},"
                "'Data SHOPEE'!${o}$2:${o}${e},{F}{r})").format(
            c=col, e=SE, o=SP_ORDER, F=F, r=r)

    def shopee_f(col, r, negate=False, qty=False):
        core  = sp_core(col, r)
        cntif = cf(r)
        inner = 'IF({c}>1,{s}/{c},{s})'.format(c=cntif, s=core)
        wrap  = 'IF({aj}{r}>0,{inner},0)'.format(aj=AJ, r=r, inner=inner)
        if qty: wrap += '*{q}{r}'.format(q=QT, r=r)
        return '=' + ('-' if negate else '') + wrap

    def sp_pay_two_types(typ1, typ2, r):
        parts1, parts2 = [], []
        for sn in SP_SHEETS:
            mr = sp_mr.get(sn, 500)
            tmpl = ("SUMIFS('{s}'!$F$2:$F${m},'{s}'!$B$2:$B${m},{F}{r},"
                    "'{s}'!$E$2:$E${m},{t},'{s}'!$A$2:$A${m},{Y}{r})")
            parts1.append(tmpl.format(s=sn, m=mr, F=F, r=r, t=typ1, Y=Y))
            parts2.append(tmpl.format(s=sn, m=mr, F=F, r=r, t=typ2, Y=Y))
        total = '(' + '+'.join(parts1) + ')+(' + '+'.join(parts2) + ')'
        cntif = cf(r)
        calc  = 'IF({c}>1,({t})/{c},({t}))'.format(c=cntif, t=total)
        return '=IF({F}{r}="",0,{calc})'.format(F=F, r=r, calc=calc)

    def sp_pay_f(typ, r):
        parts = []
        for sn in SP_SHEETS:
            mr = sp_mr.get(sn, 500)
            parts.append(
                "SUMIFS('{s}'!$F$2:$F${m},'{s}'!$B$2:$B${m},{F}{r},"
                "'{s}'!$E$2:$E${m},{t},'{s}'!$A$2:$A${m},{Y}{r})".format(
                    s=sn, m=mr, F=F, r=r, t=typ, Y=Y))
        total = '+'.join(parts)
        cntif = cf(r)
        calc  = 'IF({c}>1,({t})/{c},({t}))'.format(c=cntif, t=total)
        return '=IF({F}{r}="",0,{calc})'.format(F=F, r=r, calc=calc)

    def ai_f(r):
        return '=IF({F}{r}="",0,{aj}{r}-SUM({aa}{r}:{ah}{r}))'.format(
            F=F, r=r, aj=L[CAJ], aa=AA, ah=AH)

    def cost_unit_f(r):
        return (
            '=IFERROR(IF(${aj}{r}>0,'
            'VLOOKUP({sku}{r},IMPORTRANGE("{url}","CostORG2!$c$3:$n$80000"),12,0)'
            ',0),"")'
        ).format(aj=L[CAJ], r=r, sku=SKU, url=COST_URL)

    def cost_total_f(r):
        return '={cak}{r}*{qt}{r}'.format(cak=L[CAK], r=r, qt=QT)

    def profit_f(r):
        return '={aj}{r}+{cal}{r}'.format(aj=L[CAJ], r=r, cal=L[CAL])

    def profit_pct_f(r):
        return '=IF({aa}{r},{cam}{r}/{aa}{r},0)'.format(
            aa=L[CAA], r=r, cam=L[CAM])

    def note_f(r):
        return (
            '=INDEX(IMPORTRANGE("{url}","2026!$I$2:$Z$15000"),'
            'MATCH(C{r},IMPORTRANGE("{url}","2026!$I$2:$I$15000"),0),18)'
        ).format(url=NOTE_URL, r=r)

    print('Writing rows...')
    for i, row_data in enumerate(rows):
        r = i + 2
        for c, val in enumerate(row_data[:N], 1):
            ws.cell(r,c).value = val
        ws.cell(r,5).value = '=INDEX(STOCK!$B:$B,MATCH(C{r},STOCK!$I:$I,0))'.format(r=r)
        sku_val = str(row_data[sku_i] or '') if sku_i is not None else ''
        if sku_val.startswith('BP051'):
            ws.cell(r,6).value = 0
        if SP_PRICE: ws.cell(r,CAA).value = shopee_f(SP_PRICE, r, negate=False, qty=True)
        if SP_AP:    ws.cell(r,CAB).value = shopee_f(SP_AP,    r, negate=True,  qty=False)
        if SP_AO:    ws.cell(r,CAC).value = shopee_f(SP_AO,    r, negate=True,  qty=False)
        if SP_AU:    ws.cell(r,CAD).value = shopee_f(SP_AU,    r, negate=True,  qty=False)
        ws.cell(r,CAE).value = sp_pay_f('"รายการปรับปรุง"', r)
        ws.cell(r,CAF).value = 0
        if SP_AS:    ws.cell(r,CAG).value = shopee_f(SP_AS,    r, negate=False, qty=False)
        if SP_AR:    ws.cell(r,CAH).value = shopee_f(SP_AR,    r, negate=False, qty=False)
        ws.cell(r,CAJ).value = sp_pay_two_types('"รายรับจากคำสั่งซื้อ"', '"รายการปรับปรุง"', r)
        ws.cell(r,CAI).value = ai_f(r)
        ws.cell(r,CAK).value = cost_unit_f(r)
        ws.cell(r,CAL).value = cost_total_f(r)
        ws.cell(r,CAM).value = profit_f(r)
        ws.cell(r,CAN).value = profit_pct_f(r)
        ws.cell(r,CAO).value = note_f(r)
        ws.cell(r,CAP).value = ''
        for fc in range(CAA, CAJ+1):
            ws.cell(r, fc).number_format = '#,##0.00'
        ws.cell(r,CAK).number_format = '#,##0.00'
        ws.cell(r,CAL).number_format = '#,##0.00'
        ws.cell(r,CAM).number_format = '#,##0.00'
        ws.cell(r,CAN).number_format = '0.00%'

    # ── #ไม่พบ Order rows ────────────────────────────────────────────────────
    print('Scanning for #ไม่พบ Order rows in SPจ่าย...')
    YELLOW_FILL = PatternFill('solid', fgColor='FFFF99')
    no_order_count = 0
    for sname in SP_SHEETS:
        if sname not in wb.sheetnames: continue
        ws_sp = wb[sname]
        for r_sp in range(2, ws_sp.max_row + 1):
            shop_qc = str(ws_sp.cell(r_sp, 9).value or '').strip()
            typ_val = str(ws_sp.cell(r_sp, 5).value or '').strip()
            amt_val = ws_sp.cell(r_sp, 6).value
            if not shop_qc.startswith('#ไม่พบ Order'): continue
            if 'ถอนเงิน' in typ_val: continue
            if amt_val is None or str(amt_val).strip() in ('', '0', 'None'): continue
            try:
                amt_float = float(amt_val)
            except:
                continue
            if amt_float == 0: continue
            shop_name = str(ws_sp.cell(r_sp, 1).value or '').strip()
            new_r = ws.max_row + 1
            ws.cell(new_r, 13).value = 'หักค่าธรรมเนียม'
            ws.cell(new_r, 25).value = shop_name
            ws.cell(new_r, CAE).value = amt_float
            ws.cell(new_r, CAJ).value = amt_float
            ws.cell(new_r, CAE).number_format = '#,##0.00'
            ws.cell(new_r, CAJ).number_format = '#,##0.00'
            for c in range(1, len(all_hdrs) + 1):
                ws.cell(new_r, c).fill = YELLOW_FILL
                ws.cell(new_r, c).alignment = Alignment(wrap_text=False, vertical='center')
            no_order_count += 1
            print(f'  เพิ่ม: {shop_name} | {typ_val} | {amt_float}')
    print(f'เพิ่มแถว #ไม่พบ Order ทั้งหมด: {no_order_count} แถว')

    # ── Formatting เงินเข้าบริษัท ────────────────────────────────────────────
    print('Formatting...')
    COST_FILL = 'FFF2CC'
    for c in range(1, len(all_hdrs)+1):
        cell = ws.cell(1,c)
        if c <= N:
            cell.fill = PatternFill('solid', fgColor='0F4267')
            cell.font = Font(color='FFFFFF', bold=True)
        elif c <= CAJ:
            cell.fill = PatternFill('solid', fgColor='E2EFDA')
            cell.font = Font(color='000000', bold=True)
        else:
            cell.fill = PatternFill('solid', fgColor=COST_FILL)
            cell.font = Font(color='000000', bold=True)
        cell.alignment = Alignment(wrap_text=False, vertical='center')
        ws.column_dimensions[get_column_letter(c)].width = 20

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='center')
    ws.freeze_panes = 'A2'

    # ── ใบปะหน้า SHOPEE ──────────────────────────────────────────────────────
    _m = _re.search(r'SP(\d{2})(\d{2})', INPUT)
    if _m:
        month_str = f"{_m.group(2)}/20{_m.group(1)}"
    else:
        _m2 = _re.search(r'SP0*(\d{1,2})(?!\d)', INPUT)
        month_str = f"{int(_m2.group(1)):02d}/2026" if _m2 else "MM/YYYY"

    if 'ใบปะหน้า SHOPEE' in wb.sheetnames: del wb['ใบปะหน้า SHOPEE']
    ws_cv = wb.create_sheet('ใบปะหน้า SHOPEE', 1)
    ws_cv.sheet_view.showGridLines = True

    SHOPS_CV = [
        ('SP-MAFIA',     'SHOPEE-MAFIA'),
        ('SP-UTOPIA',    'SHOPEE-UTOPIA'),
        ('SP-TRC',       'SHOPEE-TRC'),
        ('SP-FREEROAD',  'SHOPEE-FREEROAD'),
        ('SP-WORKFORCE', 'SHOPEE-WORKFORCE'),
    ]
    SRC_SH  = "'เงินเข้าบริษัท'"
    NUM_FMT = '#,##0.00'
    SC = 4; TC = 9

    def _sf(col, sn): return f'=SUMIF({SRC_SH}!${Y}:${Y},"{sn}",{SRC_SH}!${col}:${col})'
    def _sd(c='000000'): return Side(style='thin', color=c)
    BRD   = Border(left=_sd(), right=_sd(), top=_sd(), bottom=_sd())
    BRD_N = Border()

    def _f(h): return PatternFill('solid', fgColor=h)
    F_WHITE=_f('FFFFFF'); F_NAVY=_f('1F3864'); F_BLUE_A=_f('9DC3E6')
    F_RED=_f('CC0000');   F_PEACH=_f('FAC090')

    def _fn(b=False, col='000000', sz=11):
        return Font(bold=b, color=col, size=sz, name='Calibri')
    FN_TITLE=_fn(b=True,col='1F3864',sz=14); FN_HDR=_fn(b=True,col='FFFFFF',sz=11)
    FN_CAT_EX=_fn(b=True,col='FFFFFF',sz=10); FN_CAT_IN=_fn(b=True,col='1F3864',sz=10)
    FN_BOLD=_fn(b=True); FN_NORM=_fn()
    FN_SML=Font(color='555555',size=9,name='Calibri',italic=True)
    AL_L=Alignment(horizontal='left',vertical='center')
    AL_C=Alignment(horizontal='center',vertical='center',wrap_text=True)
    AL_R=Alignment(horizontal='right',vertical='center')

    ws_cv.column_dimensions['A'].width=13; ws_cv.column_dimensions['B'].width=33
    ws_cv.column_dimensions['C'].width=12
    for lt in ['D','E','F','G','H','I']: ws_cv.column_dimensions[lt].width=15

    def _rh(r,h): ws_cv.row_dimensions[r].height=h

    def _set(r,c,val=None,fill=None,font=None,fmt=None,align=None,brd=BRD):
        cell=ws_cv.cell(r,c)
        if val  is not None: cell.value=val
        if fill: cell.fill=fill
        if font: cell.font=font
        if fmt:  cell.number_format=fmt
        if align is not None: cell.alignment=align
        if brd  is not None: cell.border=brd
        return cell

    def fill_row(r,fill,font=None,brd=BRD):
        for c in range(1,TC+1): _set(r,c,fill=fill,font=font,brd=brd)

    def data_row(r,label,note,bg,vf,fmt=NUM_FMT,bold=False,col_a_fill=None):
        vfont=FN_BOLD if bold else FN_NORM
        ca_fill=col_a_fill if col_a_fill else bg
        _set(r,1,fill=ca_fill,brd=BRD); _set(r,2,label,bg,vfont,align=AL_L)
        _set(r,3,note,bg,FN_SML,align=AL_R)
        for i,(_,sn) in enumerate(SHOPS_CV,SC): _set(r,i,vf(sn),bg,vfont,fmt,AL_R)
        sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
        _set(r,TC,f'=SUM({sh0}{r}:{sh1}{r})',bg,_fn(b=True),fmt,AL_R)

    _rh(1,34); fill_row(1,F_WHITE,brd=BRD_N)
    _set(1,1,f'สรุปยอดขาย SHOPEE เดือน {month_str}',F_WHITE,FN_TITLE,
         align=Alignment(horizontal='center',vertical='center'),brd=BRD_N)
    ws_cv.merge_cells(start_row=1,start_column=1,end_row=1,end_column=TC)

    _rh(2,28)
    for c,h in enumerate(['หมวด','รายการ','หมายเหตุ']+[s[0] for s in SHOPS_CV]+['รวม'],1):
        _set(2,c,h,F_NAVY,FN_HDR,align=AL_C)

    _rh(3,24)
    data_row(3,'ยอดขาย','',F_WHITE,lambda sn:_sf(L[CAA],sn),bold=True,col_a_fill=F_BLUE_A)
    sales_row=3
    ws_cv.cell(3,1).value='รายได้'; ws_cv.cell(3,1).font=FN_CAT_IN; ws_cv.cell(3,1).alignment=AL_C

    exp_items=[
        ('ค่าธรรมเนียมการโอน',         L[CAB]),
        ('ค่าคอมมิชชั่น',              L[CAC]),
        ('ค่าบริการ',                   L[CAD]),
        ('รายการปรับปรุงจาก SHOPEE',    L[CAE]),
        ('ค่าธรรมเนียมการจ่ายนอกรอบ',  L[CAF]),
        ('ค่าจัดส่ง Shopee ออกให้',    L[CAG]),
        ('ค่าจัดส่งที่ชำระโดยผู้ซื้อ', L[CAH]),
        ('ค่าจัดส่งที่ชำระโดยผู้ขาย',  L[CAI]),
    ]
    for idx,(label,col) in enumerate(exp_items):
        r=4+idx; _rh(r,20)
        data_row(r,label,'',F_WHITE,lambda sn,c=col:_sf(c,sn),col_a_fill=F_RED)

    ws_cv.merge_cells(start_row=4,start_column=1,end_row=11,end_column=1)
    c_cat=ws_cv.cell(4,1); c_cat.value='ค่าใช้จ่าย\nSHOPEE'; c_cat.fill=F_RED
    c_cat.font=FN_CAT_EX; c_cat.alignment=AL_C; c_cat.border=BRD
    for r in range(5,12): ws_cv.cell(r,1).fill=F_RED; ws_cv.cell(r,1).border=BRD

    _rh(12,24); fill_row(12,F_PEACH)
    _set(12,1,fill=F_PEACH,brd=BRD); _set(12,2,'จำนวนเงินเข้าบริษัท',F_PEACH,FN_BOLD,align=AL_L)
    _set(12,3,'',F_PEACH,brd=BRD)
    for ci in range(SC,TC):
        cl=get_column_letter(ci); _set(12,ci,f'=SUM({cl}3:{cl}11)',F_PEACH,FN_BOLD,NUM_FMT,AL_R)
    sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
    _set(12,TC,f'=SUM({sh0}12:{sh1}12)',F_PEACH,FN_BOLD,NUM_FMT,AL_R)
    income_row=12

    _rh(13,20)
    data_row(13,'ต้นทุนค่าสินค้า','',F_WHITE,lambda sn:_sf(L[CAL],sn))

    adj_labels=['ปรับค่าไร้โอตอล','เงินเข้าค่าธรรมเนียมเข้ารับ (freeroad)','หักค่าธรรมเนียมเข้ารับ']
    for i,label in enumerate(adj_labels):
        r=14+i; _rh(r,18)
        _set(r,1,fill=F_WHITE,brd=BRD); _set(r,2,label,F_WHITE,FN_NORM,align=AL_L)
        _set(r,3,'',F_WHITE,brd=BRD)
        for ci in range(SC,TC): _set(r,ci,0,F_WHITE,FN_NORM,NUM_FMT,AL_R)
        sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
        _set(r,TC,f'=SUM({sh0}{r}:{sh1}{r})',F_WHITE,FN_BOLD,NUM_FMT,AL_R)

    _rh(17,26); fill_row(17,F_BLUE_A)
    _set(17,1,fill=F_BLUE_A,brd=BRD); _set(17,2,'กำไรเบื้องต้น (บาท)',F_BLUE_A,FN_BOLD,align=AL_L)
    _set(17,3,'',F_BLUE_A,brd=BRD)
    for ci in range(SC,TC):
        cl=get_column_letter(ci); _set(17,ci,f'=SUM({cl}12:{cl}16)',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)
    sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
    _set(17,TC,f'=SUM({sh0}17:{sh1}17)',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)
    gross_row=17

    for r,(label,rate) in [(18,('ต้นทุนค่ากล่อง',-3.785)),(19,('ต้นทุนค่า PACKING',-7.845))]:
        _rh(r,20); _set(r,1,fill=F_WHITE,brd=BRD); _set(r,2,label,F_WHITE,FN_NORM,align=AL_L)
        _set(r,3,None,F_WHITE,brd=BRD)
        for ci in range(SC,TC):
            cl=get_column_letter(ci); _set(r,ci,f'=$C${r}*{cl}17',F_WHITE,FN_NORM,NUM_FMT,AL_R)
        sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
        _set(r,TC,f'=SUM({sh0}{r}:{sh1}{r})',F_WHITE,FN_BOLD,NUM_FMT,AL_R)

    _rh(20,26); fill_row(20,F_PEACH)
    _set(20,1,'กำไรสุทธิ',F_PEACH,FN_BOLD,align=AL_C)
    ws_cv.merge_cells(start_row=20,start_column=1,end_row=20,end_column=3)
    for ci in range(SC,TC):
        cl=get_column_letter(ci); _set(20,ci,f'=SUM({cl}17:{cl}19)',F_PEACH,FN_BOLD,NUM_FMT,AL_R)
    sh0=get_column_letter(SC); sh1=get_column_letter(TC-1)
    _set(20,TC,f'=SUM({sh0}20:{sh1}20)',F_PEACH,FN_BOLD,NUM_FMT,AL_R)
    net_row=20

    _rh(21,24); fill_row(21,F_PEACH)
    _set(21,1,'กำไร (%)',F_PEACH,FN_BOLD,align=AL_C)
    ws_cv.merge_cells(start_row=21,start_column=1,end_row=21,end_column=3)
    for ci in range(SC,TC+1):
        cl=scl=get_column_letter(ci)
        _set(21,ci,f'=IF({scl}{sales_row}<>0,{cl}{net_row}/{scl}{sales_row},0)',
             F_PEACH,FN_BOLD,'0.00%',AL_R)

    ws_cv.column_dimensions['J'].width=18; ws_cv.column_dimensions['K'].width=18
    COM_HDR=25; COM_BASE=26; COM_P1=27; COM_P2=28; COM_P3=29; COM_SUM=30
    _rh(COM_HDR,28); _rh(COM_BASE,22); _rh(COM_P1,20); _rh(COM_P2,20); _rh(COM_P3,20); _rh(COM_SUM,24)

    FN_BLK_BOLD=Font(color='000000',bold=True,size=11,name='Calibri')
    FN_BLK_SML=Font(color='000000',size=9,name='Calibri',italic=True)
    com_hdrs=['ร้านค้า','','%','SP-MAFIA','SP-UTOPIA','SP-TRC','SP-FREEROAD','SP-WORKFORCE',
              'หักสินค้าขาดทุน','คืน/หัก\nสูญหาย/ชำรุด','ค่าคอมสุทธิ']
    for ci,h in enumerate(com_hdrs,1):
        _set(COM_HDR,ci,h,F_PEACH,FN_BLK_BOLD,
             align=Alignment(horizontal='center',vertical='center',wrap_text=True))

    _set(COM_BASE,1,'กำไร(ยอดเบิกค่าคอม)',F_PEACH,FN_BLK_BOLD,align=AL_L)
    ws_cv.merge_cells(f'A{COM_BASE}:B{COM_BASE}')
    _set(COM_BASE,3,'ค่าคอม',F_PEACH,FN_BLK_SML,align=AL_C)
    for ci in range(SC,TC):
        cl=get_column_letter(ci); _set(COM_BASE,ci,f'={cl}{net_row}',F_PEACH,FN_BLK_BOLD,NUM_FMT,AL_R)
    for ci in [9,10,11]: _set(COM_BASE,ci,None,F_PEACH,brd=BRD)

    people=[('แอร์',0.10),('เอ',0.02),('พีพี',0)]
    for idx,(name,rate) in enumerate(people):
        r=COM_P1+idx; _set(r,1,name,F_WHITE,FN_NORM,align=AL_L)
        ws_cv.merge_cells(f'A{r}:B{r}')
        if rate>0:
            _set(r,3,rate,F_WHITE,FN_SML,fmt='0%',align=AL_C)
            for ci in range(SC,TC):
                cl=get_column_letter(ci); _set(r,ci,f'=$C${r}*{cl}{COM_BASE}',F_WHITE,FN_NORM,NUM_FMT,AL_R)
        else:
            _set(r,3,0,F_WHITE,FN_SML,fmt='0',align=AL_C)
            for ci in range(SC,TC): _set(r,ci,0,F_WHITE,FN_NORM,NUM_FMT,AL_R)
        _set(r,9,0,F_WHITE,FN_NORM,NUM_FMT,AL_R); _set(r,10,0,F_WHITE,FN_NORM,NUM_FMT,AL_R)
        _set(r,11,f'=SUM(D{r}:H{r})+I{r}+J{r}',F_WHITE,FN_BOLD,NUM_FMT,AL_R)

    _set(COM_SUM,1,'รวม',F_BLUE_A,FN_BOLD,align=AL_C)
    _set(COM_SUM,2,None,F_BLUE_A,brd=BRD); _set(COM_SUM,3,None,F_BLUE_A,brd=BRD)
    for ci in range(SC,TC):
        cl=get_column_letter(ci)
        _set(COM_SUM,ci,f'=SUM({cl}{COM_P1}:{cl}{COM_P3})',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)
    _set(COM_SUM,9,f'=SUM(I{COM_P1}:I{COM_P3})',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)
    _set(COM_SUM,10,f'=SUM(J{COM_P1}:J{COM_P3})',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)
    _set(COM_SUM,11,f'=SUM(K{COM_P1}:K{COM_P3})',F_BLUE_A,FN_BOLD,NUM_FMT,AL_R)

    ws_cv.print_area='A1:K30'
    ws_cv.page_setup.orientation='landscape'; ws_cv.page_setup.fitToPage=True
    ws_cv.page_setup.fitToWidth=1; ws_cv.page_setup.fitToHeight=0
    ws_cv.page_margins=PageMargins(left=0.5,right=0.5,top=0.6,bottom=0.6)

    print(f'ใบปะหน้า SHOPEE → {month_str}')
    wb.save(OUTPUT)
    print(f'\nDone: {len(rows)} rows saved -> {OUTPUT}')


if __name__ == '__main__':
    INPUT  = sys.argv[1]
    OUTPUT = sys.argv[2]
    run(INPUT, OUTPUT)
