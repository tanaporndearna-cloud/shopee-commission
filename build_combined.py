# -*- coding: utf-8 -*-
"""
build_combined.py  –  สร้าง input_combined.xlsx (10 ชีต)
ใช้ได้ทั้งแบบ import (run()) และ command-line (python build_combined.py BASE OUT [STOCK_CSV])
"""
import sys, os, glob, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


def run(BASE, OUT, STOCK_CSV=None):
    def safe_read(path, **kw):
        try:
            return pd.read_excel(path, **kw)
        except Exception as e:
            print(f'  [warn] {os.path.basename(path)}: {e}')
            return pd.DataFrame()

    # ── STOCK ──────────────────────────────────────────────────────────────────
    if STOCK_CSV and os.path.exists(STOCK_CSV):
        df_stock = pd.read_csv(STOCK_CSV, dtype=str)
        print(f'STOCK (CSV): {len(df_stock)} rows')
    else:
        stock_path = os.path.join(BASE, 'STOCK', 'STOCK.xlsx')
        if not os.path.exists(stock_path):
            stock_path = os.path.join(BASE, 'STOCK.xlsx')
        df_stock = safe_read(stock_path) if os.path.exists(stock_path) else pd.DataFrame()
        print(f'STOCK: {len(df_stock)} rows')

    order_to_abb    = {}
    abb_to_order    = {}
    order_to_shopee = {}

    for _, row in df_stock.iterrows():
        ono  = str(row.iloc[0]).strip()  if pd.notna(row.iloc[0])  else ''
        abb  = str(row.iloc[8]).strip()  if len(row) > 8  and pd.notna(row.iloc[8])  else ''
        shop = str(row.iloc[20]).strip() if len(row) > 20 and pd.notna(row.iloc[20]) else ''
        if ono in ('nan','0','') or not abb: continue
        if ono not in order_to_abb: order_to_abb[ono] = abb
        abb_to_order[abb] = ono
        if shop.startswith('SHOPEE-'): order_to_shopee[ono] = shop

    print(f'  order_to_abb: {len(order_to_abb)} | order_to_shopee: {len(order_to_shopee)}')

    # ── Data ERP ───────────────────────────────────────────────────────────────
    df_erp_raw = safe_read(os.path.join(BASE, 'Data ERP', 'Data ERP.xlsx'))
    print(f'Data ERP: {len(df_erp_raw)} rows')

    def build_abb_shop_map(df):
        m = {}
        if df.empty or len(df.columns) < 23: return m
        abb_col  = df.columns[2]
        shop_col = 'พนักงานขาย' if 'พนักงานขาย' in df.columns else df.columns[22]
        for _, row in df.iterrows():
            abb  = str(row[abb_col]).strip()  if pd.notna(row[abb_col])  else ''
            shop = str(row[shop_col]).strip() if pd.notna(row[shop_col]) else ''
            if abb and shop and not abb.startswith('CA'): m[abb] = shop
        return m

    def build_order_shop_map(df):
        m = {}
        if df.empty or len(df.columns) < 23: return m
        abb_col  = df.columns[2]
        shop_col = 'พนักงานขาย' if 'พนักงานขาย' in df.columns else df.columns[22]
        for _, row in df.iterrows():
            abb  = str(row[abb_col]).strip()  if pd.notna(row[abb_col])  else ''
            shop = str(row[shop_col]).strip() if pd.notna(row[shop_col]) else ''
            if not abb or not shop or abb.startswith('CA'): continue
            ono = abb_to_order.get(abb, '')
            if ono: m[ono] = shop
        return m

    erp_abb_shop   = build_abb_shop_map(df_erp_raw)
    erp_order_shop = build_order_shop_map(df_erp_raw)
    erp_abb_set    = set(str(r.iloc[2]).strip() for _, r in df_erp_raw.iterrows()
                         if len(r) > 2 and pd.notna(r.iloc[2]) and not str(r.iloc[2]).strip().startswith('CA'))
    print(f'  erp_abb_set: {len(erp_abb_set)} ABBs')

    # ── ABBHSP ─────────────────────────────────────────────────────────────────
    abbhsp_dfs         = {}
    abbhsp_abb_shop    = {}
    abbhsp_order_shop  = {}
    abbhsp_rows_by_abb = {}

    for f in sorted(glob.glob(os.path.join(BASE, 'ABBHSP', '*.xlsx'))):
        if os.path.basename(f).startswith('~$'): continue
        name = os.path.splitext(os.path.basename(f))[0]
        df = safe_read(f)
        abbhsp_dfs[name] = df
        abbhsp_abb_shop.update(build_abb_shop_map(df))
        abbhsp_order_shop.update(build_order_shop_map(df))
        for _, row in df.iterrows():
            abb = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
            if abb: abbhsp_rows_by_abb.setdefault(abb, []).append(row)
        print(f'{name}: {len(df)} rows')

    abbhsp_rows_by_order = {}
    for abb, rows in abbhsp_rows_by_abb.items():
        ono = abb_to_order.get(abb, '')
        if ono: abbhsp_rows_by_order.setdefault(ono, []).extend(rows)

    # ── Data SHOPEE ────────────────────────────────────────────────────────────
    SHOP_PREFIXES = {
        'SP-MAFIA':     'SHOPEE-MAFIA',
        'SP-TRC':       'SHOPEE-TRC',
        'SP-UTOPIA':    'SHOPEE-UTOPIA',
        'SP-UTOIA':     'SHOPEE-UTOPIA',
        'SP-FREEROAD':  'SHOPEE-FREEROAD',
        'SP-WORKFORCE': 'SHOPEE-WORKFORCE',
    }
    sp_frames = []
    shopee_dir = os.path.join(BASE, 'Data SHOPEE')
    for f in sorted(glob.glob(os.path.join(shopee_dir, '*.xlsx'))):
        if os.path.basename(f).startswith('~$'): continue
        fname_upper = os.path.basename(f).upper()
        shop = None
        for prefix, shopname in SHOP_PREFIXES.items():
            if fname_upper.startswith(prefix.upper()):
                shop = shopname; break
        if shop is None: continue
        df = safe_read(f)
        if df.empty: continue
        df.insert(0, 'ชื่อร้าน', shop)
        sp_frames.append(df)
        print(f'  SHOPEE {os.path.basename(f)}: {len(df)} rows -> {shop}')
    for subfolder, shop in SHOP_PREFIXES.items():
        fp = os.path.join(shopee_dir, subfolder)
        if not os.path.isdir(fp): continue
        for f in sorted(glob.glob(os.path.join(fp, '*.xlsx'))):
            if os.path.basename(f).startswith('~$'): continue
            df = safe_read(f)
            if df.empty: continue
            df.insert(0, 'ชื่อร้าน', shop)
            sp_frames.append(df)

    df_shopee = pd.concat(sp_frames, ignore_index=True) if sp_frames else pd.DataFrame()
    print(f'Data SHOPEE รวม: {len(df_shopee)} rows')

    # ── การเงิน SHOPEE ─────────────────────────────────────────────────────────
    FIN_PREFIXES = [
        ('การเงิน-MAFIA',     'SHOPEE-MAFIA',     'SPจ่ายMAFIA'),
        ('การเงิน-TRC',       'SHOPEE-TRC',       'SPจ่ายTRC'),
        ('การเงิน -TRC',      'SHOPEE-TRC',       'SPจ่ายTRC'),
        ('การเงิน-UTOPIA',    'SHOPEE-UTOPIA',    'SPจ่ายUTOPIA'),
        ('การเงิน-FREEROAD',  'SHOPEE-FREEROAD',  'SPจ่ายFREEROAD'),
        ('การเงิน-WORKFORCE', 'SHOPEE-WORKFORCE', 'SPจ่ายWORKFORCE'),
    ]
    fin_raw_files = {}
    fin_dir = None
    for fin_folder_name in ['การเงิน SHOPEE', 'การเงิน']:
        candidate = os.path.join(BASE, fin_folder_name)
        if os.path.isdir(candidate):
            fin_dir = candidate
            print(f'ใช้โฟลเดอร์การเงิน: {fin_folder_name}')
            break

    if fin_dir:
        for f in sorted(glob.glob(os.path.join(fin_dir, '*.xlsx'))):
            if os.path.basename(f).startswith('~$'): continue
            fname = os.path.basename(f)
            for prefix, shop, sheet_name in FIN_PREFIXES:
                if fname.startswith(prefix):
                    fin_raw_files.setdefault(sheet_name, []).append((shop, f))
                    print(f'  การเงิน {fname} -> {sheet_name}')
                    break
        FIN_MAP = {
            'การเงิน-MAFIA':     ('SHOPEE-MAFIA',     'SPจ่ายMAFIA'),
            'การเงิน-TRC':       ('SHOPEE-TRC',       'SPจ่ายTRC'),
            'การเงิน-UTOPIA':    ('SHOPEE-UTOPIA',    'SPจ่ายUTOPIA'),
            'การเงิน-FREEROAD':  ('SHOPEE-FREEROAD',  'SPจ่ายFREEROAD'),
            'การเงิน-WORKFORCE': ('SHOPEE-WORKFORCE', 'SPจ่ายWORKFORCE'),
        }
        for subfolder, (shop, sheet_name) in FIN_MAP.items():
            fp = os.path.join(fin_dir, subfolder)
            if not os.path.isdir(fp): continue
            for f in sorted(glob.glob(os.path.join(fp, '*.xlsx'))):
                if os.path.basename(f).startswith('~$'): continue
                fin_raw_files.setdefault(sheet_name, []).append((shop, f))

    fin_raws          = {}
    missing_order_nos = set()
    for sheet_name, file_list in fin_raw_files.items():
        frames = []
        shop = file_list[0][0]
        for s, fpath in file_list:
            df = safe_read(fpath, header=17)
            if df.empty: continue
            cols = list(df.columns)
            rmap = {}
            for i, n in enumerate(['วันที่','ประเภทการทำธุรกรรม','คำอธิบาย','รหัสคำสั่งซื้อ','รูปแบบธุรกรรม','จำนวนเงิน']):
                if i < len(cols): rmap[cols[i]] = n
            df = df.rename(columns=rmap)
            if 'รหัสคำสั่งซื้อ' not in df.columns: continue
            df = df[df['รหัสคำสั่งซื้อ'].notna()]
            frames.append(df)
        if not frames: continue
        raw = pd.concat(frames, ignore_index=True)
        fin_raws[sheet_name] = (shop, raw)
        for _, row in raw.iterrows():
            order_no = str(row.get('รหัสคำสั่งซื้อ','')).strip()
            typ      = str(row.get('ประเภทการทำธุรกรรม','')).strip()
            if 'ถอนเงิน' in typ: continue
            abb = order_to_abb.get(order_no, '')
            if abb and abb not in erp_abb_set and order_no in abbhsp_rows_by_order:
                missing_order_nos.add(order_no)

    print(f'\norder ที่ Col I จะ Error แต่มีใน ABBHSP: {len(missing_order_nos)}')

    extra_rows = []
    for ono in sorted(missing_order_nos):
        rows = abbhsp_rows_by_order.get(ono, [])
        extra_rows.extend(rows)
        print(f'  {ono}: {len(rows)} row(s) appended')

    extra_erp_start_idx = len(df_erp_raw)
    df_erp_extended = df_erp_raw.copy()
    if extra_rows:
        df_extra = pd.DataFrame(extra_rows).reset_index(drop=True)
        df_erp_extended = pd.concat([df_erp_raw, df_extra], ignore_index=True)
        erp_abb_shop.update(build_abb_shop_map(df_extra))
        erp_order_shop.update(build_order_shop_map(df_extra))

    df_erp_final = df_erp_extended.copy()
    df_erp_final.insert(4, 'เลขสั่งซื้อ Shopee',
        df_erp_final.iloc[:, 2].map(lambda x: 0 if str(x).strip().startswith('CA') else abb_to_order.get(str(x).strip(), '')))
    df_erp_final.insert(5, 'เลขสั่งซื้อ Shopee (copy)', df_erp_final.iloc[:, 4])
    print(f'Data ERP หลังเพิ่ม: {len(df_erp_final)} rows')

    fin_sheets  = {}
    fin_wd_rows = {}
    for sheet_name, (shop, raw) in fin_raws.items():
        out_rows   = []
        wd_indices = set()
        for _, row in raw.iterrows():
            order_no = str(row.get('รหัสคำสั่งซื้อ','')).strip()
            typ      = str(row.get('ประเภทการทำธุรกรรม','')).strip()
            amt      = row.get('จำนวนเงิน','')
            date_val = row.get('วันที่','')
            is_wd    = 'ถอนเงิน' in typ
            abb      = '' if is_wd else order_to_abb.get(order_no,'')
            if is_wd and amt != '' and pd.notna(amt):
                try:    col_g = abs(float(amt))
                except: col_g = amt
            else:
                col_g = ''
            if is_wd and date_val != '' and pd.notna(date_val):
                try:
                    d = str(date_val).strip()
                    d = re.split(r'[ T]', d)[0]
                    parts = d.split('-')
                    col_h = f'{int(parts[2])}/{int(parts[1])}/{parts[0]}' if len(parts)==3 else d
                except:
                    col_h = str(date_val)
            else:
                col_h = ''
            if is_wd:
                col_i = ''
            else:
                col_i = erp_order_shop.get(order_no, '') or abbhsp_order_shop.get(order_no, '') or \
                        order_to_shopee.get(order_no, f'#ไม่พบ Order:{order_no}')
            if is_wd: wd_indices.add(len(out_rows))
            out_rows.append({
                'ชื่อร้านค้า': shop, 'รหัสคำสั่งซื้อ': order_no, 'ABB': abb,
                'วันที่': date_val, 'ประเภทการทำธุรกรรม': typ, 'จำนวนเงิน': amt,
                'ยอดการถอนเงิน': col_g, 'วันที่ถอนเงิน': col_h, 'Shop QC': col_i,
            })
        fin_sheets[sheet_name]  = pd.DataFrame(out_rows)
        fin_wd_rows[sheet_name] = wd_indices
        errors = sum(1 for r in out_rows if str(r['Shop QC']).startswith('#'))
        print(f'{sheet_name}: {len(out_rows)} rows ({len(wd_indices)} ถอนเงิน, {errors} error)')

    mismatch_correct = {}
    for sheet_name, df in fin_sheets.items():
        for _, row in df.iterrows():
            col_a    = str(row.get('ชื่อร้านค้า','')).strip()
            col_i    = str(row.get('Shop QC','')).strip()
            order_no = str(row.get('รหัสคำสั่งซื้อ','')).strip()
            if col_i and not col_i.startswith('#') and col_a != col_i:
                mismatch_correct[order_no] = col_a

    print(f'\nMismatch Col A vs Col I: {len(mismatch_correct)} orders')
    erp_mismatch_rows = set()
    shopee_col = 'เลขสั่งซื้อ Shopee'
    for i, row in df_erp_final.iterrows():
        ono = str(row.get(shopee_col, '')).strip()
        if ono in mismatch_correct:
            if 'พนักงานขาย' in df_erp_final.columns:
                df_erp_final.at[i, 'พนักงานขาย'] = mismatch_correct[ono]
            erp_mismatch_rows.add(i)
    print(f'แก้ใน Data ERP: {len(erp_mismatch_rows)} แถว')

    # ── เขียน Excel ────────────────────────────────────────────────────────────
    with pd.ExcelWriter(OUT, engine='openpyxl') as writer:
        if not df_stock.empty:
            df_stock.to_excel(writer, sheet_name='STOCK', index=False)
        else:
            pd.DataFrame({'(ไม่มีไฟล์ STOCK)':[]}).to_excel(writer, sheet_name='STOCK', index=False)
        df_erp_final.to_excel(writer, sheet_name='Data ERP', index=False)
        for name, df in abbhsp_dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)
        df_shopee.to_excel(writer, sheet_name='Data SHOPEE', index=False)
        for sname, df in fin_sheets.items():
            df.to_excel(writer, sheet_name=sname, index=False)

    # ── Formatting ─────────────────────────────────────────────────────────────
    WD_FILL       = PatternFill('solid', fgColor='FFD1AD')
    XTRA_FILL     = PatternFill('solid', fgColor='FFA9A9')
    MISMATCH_FILL = PatternFill('solid', fgColor='CFE0EB')
    wb = load_workbook(OUT)
    ws_erp = wb['Data ERP']
    max_col_erp = ws_erp.max_column
    if extra_rows:
        for i in range(len(extra_rows)):
            excel_row = extra_erp_start_idx + i + 2
            for col in range(1, max_col_erp + 1):
                ws_erp.cell(excel_row, col).fill = XTRA_FILL
    for df_idx in erp_mismatch_rows:
        excel_row = df_idx + 2
        for col in range(1, max_col_erp + 1):
            ws_erp.cell(excel_row, col).fill = MISMATCH_FILL
    for sheet_name, wd_indices in fin_wd_rows.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        for di in wd_indices:
            for col in range(1, ws.max_column + 1):
                ws.cell(di + 2, col).fill = WD_FILL

    if not df_stock.empty:
        for r in range(2, ws_erp.max_row + 1):
            ws_erp.cell(r, 5).value = f'=IFERROR(INDEX(STOCK!$B:$B,MATCH(C{r},STOCK!$I:$I,0)),"")'         
           
    FIN_SHEETS_F = ['SPจ่ายMAFIA','SPจ่ายTRC','SPจ่ายUTOPIA','SPจ่ายFREEROAD','SPจ่ายWORKFORCE']
    for sname in FIN_SHEETS_F:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        for r in range(2, ws.max_row + 1):
            typ_val   = ws.cell(r, 5).value
            order_val = ws.cell(r, 2).value
            is_wd     = typ_val and 'ถอนเงิน' in str(typ_val)
            has_order = order_val and str(order_val).strip() not in ('', 'None', '-')
            if is_wd:
                ws.cell(r, 7).value = f'=ABS(F{r})'
                ws.cell(r, 7).number_format = '#,##0.00'
            if not is_wd and has_order:
                if not df_stock.empty:
                    ws.cell(r, 3).value = f'=VLOOKUP(B{r},STOCK!B:I,8,0)'
                ws.cell(r, 9).value = f"=VLOOKUP($C{r},'Data ERP'!$C$2:Y$530,23,0)"

    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical='center')

    SUMM_ORANGE = PatternFill('solid', fgColor='FF9967')
    SUMM_PINK   = PatternFill('solid', fgColor='FFCCCC')
    SUMM_FONT   = Font(bold=True)
    SUMM_ROWS = [
        ('รายรับจากคำสั่งซื้อ',                     SUMM_ORANGE),
        ('การถอนเงิน',                                SUMM_ORANGE),
        ('รายการปรับปรุง',                            SUMM_ORANGE),
        ('รายการปรับปรุงค่าขนส่งเข้ารับจาก Shopee', SUMM_PINK),
    ]
    for sname in FIN_SHEETS_F:
        if sname not in wb.sheetnames: continue
        ws      = wb[sname]
        last_dr = ws.max_row
        start_r = last_dr + 2
        for i, (label, fill) in enumerate(SUMM_ROWS):
            r = start_r + i
            ws.cell(r, 4).value = label; ws.cell(r, 4).font = SUMM_FONT; ws.cell(r, 4).fill = fill
            ws.cell(r, 5).value = 'ผลรวม'; ws.cell(r, 5).font = SUMM_FONT; ws.cell(r, 5).fill = fill
            ws.cell(r, 6).value = f'=SUMIFS($F$2:$F${last_dr},$E$2:$E${last_dr},D{r})'
            ws.cell(r, 6).font = SUMM_FONT; ws.cell(r, 6).fill = fill
            ws.cell(r, 6).number_format = '#,##0.00'

    NO_BORDER  = Border()
    WHITE_FONT = Font(color='FFFFFF', bold=True)
    BOLD_FONT  = Font(bold=True)
    FIN_HDR = {(1,3): PatternFill('solid',fgColor='FF9967'),
               (4,6): PatternFill('solid',fgColor='FCCF55'),
               (7,9): PatternFill('solid',fgColor='9379C2')}
    for sname in FIN_SHEETS_F:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.border = NO_BORDER; cell.font = BOLD_FONT
            for (c1,c2), fill in FIN_HDR.items():
                if c1 <= col <= c2: cell.fill = fill; break
    for sname in ['Data SHOPEE','Data ERP']:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        fill = PatternFill('solid', fgColor='0F4267')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.fill = fill; cell.font = WHITE_FONT; cell.border = NO_BORDER

    wb.save(OUT)
    print(f'\n✅ Saved → {OUT}')


if __name__ == '__main__':
    BASE      = sys.argv[1]
    OUT       = sys.argv[2]
    STOCK_CSV = sys.argv[3] if len(sys.argv) > 3 else None
    run(BASE, OUT, STOCK_CSV)
